import os 
import pandas as pd
import openpyxl
import xlrd
#from openpyxl import load_workbook

base_dir = "data"
files = os.listdir(base_dir)

start_col_ = 0
start_row_ = 0
count = 0

computing_start_col_ = 0
computing_start_row_ = 0
computing_count = 0

networking_start_col_ = 0
networking_start_row_ = 0
networking_count = 0


multimedia_start_col_ = 0
multimedia_start_row_ = 0
multimedia_count = 0

computing_networking_start_col_ = 0
computing_networking_start_row_ = 0
computing_networking_count = 0

computing_multimedia_start_col_ = 0
computing_multimedia_start_row_ = 0
computing_multimedia_count = 0

computing_networking_multimedia_start_col_ = 0
computing_networking_multimedia_start_row_ = 0
computing_networking_multimedia_count = 0

writer = pd.ExcelWriter('results_summary.xlsx', engine='xlsxwriter')
writer2 = pd.ExcelWriter('module_wise_summary.xlsx', engine='xlsxwriter')
writer3 = pd.ExcelWriter('specification_wise_summary.xlsx', engine='xlsxwriter')

#wb2 = load_workbook('trial.xlsx')

#del wb2['Sheet1']

for file in files:
    # get data
    print(file)
    file_path = os.path.join(base_dir,file)
    marks = pd.read_excel(file_path,skiprows=9)
    details = pd.read_excel(file_path,nrows=8,header=None,usecols="A,C")
    
    # calculate grade frequencies and grade averages
    GRADES = ("A", "B", "C", "D", "E", "F")
    grade_frequencies = {}
    grade_avgs = {}
    for grade in GRADES:
        try:
            d = marks[marks["Final Grade"] == grade]
            grade_frequencies[grade] = d["Final Grade"].count()
        except:
            d = marks[marks["Module Grade"] == grade]
            grade_frequencies[grade] = d["Module Grade"].count()
        try:
            grade_avgs[grade] = float("%.2f" % d["Final Marks"].mean())
        except:
            grade_avgs[grade] = float("%.2f" % d["Module Mark"].mean())
    
    # calculate grade percents
    grade_percents = {}
    total = sum(grade_frequencies.values())
    for g in GRADES:
        grade_percents[g] = float("%.2f" % ((grade_frequencies[g] / total) * 100))
    
    # marks details
    marks_details = {"#": grade_frequencies,
                     "%": grade_percents,
                     "Avg Mark": grade_avgs}
    df = pd.DataFrame(marks_details)
    
    # other details
    other_details = {'#': {"Students Countable": df["#"].sum(),
                           "Intermission": 0,
                           "Withdraw": 0,
                           "Course Transfer": 0},
                    '%': {},
                    'Avg Mark':{}}
    
    df2 = pd.DataFrame(other_details)
    df2.reindex(['Students Countable', 'Intermission', 'Withdraw', 'Course Transfer'])
    df3 = df.append(df2)
    df3.reindex(['A', 'B', 'C', 'D', 'E', 'F', 
                 'Students Countable', 'Intermission', 'Withdraw', 'Course Transfer'])
    
    #write to excel
    try:
        summary = {'#': {"Students Total": 0,
                        "Module Avg Mark": float("%.2f" % marks['Final Marks'].mean()),
                        "Pass %": float("%.2f" % (((df['#'].sum() - df.loc["F"]["#"]) / df['#'].sum()) * 100))},
                    '%': {},
                    'Avg Mark': {}}
    except:
        summary = {'#':{"Students Total": 0,
                        "Module Avg Mark": float("%.2f" % marks['Module Mark'].mean()),
                        "Pass %": float("%.2f" % (((df['#'].sum() - df.loc["F"]["#"]) / df['#'].sum()) * 100))},
                    '%': {},
                    'Avg Mark': {}}
        
    df4 = pd.DataFrame(summary)
    final_df = df3.append(df4)
    final_df = final_df.reindex(['A', 'B', 'C', 'D', 'E', 'F',
                                 'Students Countable', 'Intermission', 'Withdraw', 
                                 'Course Transfer', 'Students Total', 'Module Avg Mark', 'Pass %'])
    
    # write module details
    details.iloc[[0,1]][:].to_excel(writer, startrow=start_row_, startcol=start_col_, 
                                    header=False, index=False)
    
    #print(details.iloc[[0,1]][:])
    
    print(details.iloc[[3]][2])
    
    specList = list(details.iloc[[3]][2])
    stringSpecList=str(specList[0])
    
    stringSpecList = stringSpecList.replace("BSc. (Hons)","")
    stringSpecList = stringSpecList.replace("BSc (Hons)","")
    stringSpecList = stringSpecList.replace("Computer","")
    stringSpecList = stringSpecList.replace("and IT Security","")
    stringSpecList = stringSpecList.replace("& IT Security","")
    stringSpecList = stringSpecList.replace("Technologies","")
    stringSpecList = stringSpecList.replace("/","|")
    stringSpecList = stringSpecList.replace(" ", "")
    
    
    print(stringSpecList[0:25]+"...")
    alist=list(details.iloc[[1]][2])
    
    #print(str(alist[0]))
    
    #print(str(alist[0])[0:25]+"...")
    
    
    #wb2.create_sheet(str(alist[0])[0:25]+"...")
    
    
    
    
    #print(wb2.sheetnames)
    
    #del wb2['Sheet1']
    
    #sheet = wb2.add_sheet(str(alist[0])[0:25]+"...")
    #sheet.write(final_df)
    #listOfSheet = wb2.get_sheet_by_name('Sheet1')
    
    #print(listOfSheet)
    
    
    #print(final_df)
    #wb2.create_sheet(str(alist[0]))
    #wb2.save('trial.xlsx')
    
    #print("asd", str(details.iloc[[1,1]][:]))
    #wb2.save('trial.xlsx')
    
    #print(details.iloc[[1,1]][:])
    
    #print(details.iloc[[1]][:])
    
    #write summary in trial
    
    
    details.iloc[[0,1]][:].to_excel(writer2, str(alist[0])[0:25]+"...", header=False, index=False)
    final_df.to_excel(writer2, str(alist[0])[0:25]+"...", startrow=3)
    
    #print(final_df)
    # write summary
    
    final_df.to_excel(writer, startrow=start_row_+3, startcol=start_col_)
    
    
    #alldone
    if stringSpecList[0:25]+"..." == "Computing|Networking...":
        details.iloc[[0,1]][:].to_excel(writer3, stringSpecList[0:25]+"...", startrow=computing_networking_start_row_, startcol=computing_networking_start_col_, 
                                    header=False, index=False)
        
        final_df.to_excel(writer3, stringSpecList[0:25]+"...", startrow=computing_networking_start_row_+3, startcol=computing_networking_start_col_)
        print("C + N")
        
        
        
        computing_networking_count += 1
        if computing_networking_count % 2 == 0:
            computing_networking_start_col_ = 0 
            computing_networking_start_row_ += 19
        else:
            computing_networking_start_col_ += 5
            computing_networking_start_row_ += 0
    
    #alldone        
    elif stringSpecList[0:25]+"..." == "Networking...":
        details.iloc[[0,1]][:].to_excel(writer3, stringSpecList[0:25]+"...", startrow=networking_start_row_, startcol=networking_start_col_, 
                                    header=False, index=False)
        
        final_df.to_excel(writer3, stringSpecList[0:25]+"...", startrow=networking_start_row_+3, startcol=networking_start_col_)
        print("N")
        
        
        networking_count += 1
        if networking_count % 2 == 0:
            networking_start_col_ = 0 
            networking_start_row_ += 19
        else:
            networking_start_col_ += 5
            networking_start_row_ += 0
    
    #alldone    
    elif stringSpecList[0:25]+"..." == "Computing|Networking|Mult...":
        details.iloc[[0,1]][:].to_excel(writer3, stringSpecList[0:25]+"...", startrow=computing_networking_multimedia_start_row_, startcol=computing_networking_multimedia_start_col_, 
                                    header=False, index=False)
        
        print("C + N + M")
        final_df.to_excel(writer3, stringSpecList[0:25]+"...", startrow=computing_networking_multimedia_start_row_+3, startcol=computing_networking_multimedia_start_col_)
        
        
        
        computing_networking_multimedia_count += 1
        if computing_networking_multimedia_count % 2 == 0:
            computing_networking_multimedia_start_col_ = 0 
            computing_networking_multimedia_start_row_ += 19
        else:
            computing_networking_multimedia_start_col_ += 5
            computing_networking_multimedia_start_row_ += 0
    
    #alldone        
    elif stringSpecList[0:25]+"..." == "Computing|Multimedia...":
        details.iloc[[0,1]][:].to_excel(writer3, stringSpecList[0:25]+"...", startrow=computing_multimedia_start_row_, startcol=computing_multimedia_start_col_, 
                                    header=False, index=False)
        
        print("C + M")
        final_df.to_excel(writer3, stringSpecList[0:25]+"...", startrow=computing_multimedia_start_row_+3, startcol=computing_multimedia_start_col_)
     
        
        computing_multimedia_count += 1
        if computing_multimedia_count % 2 == 0:
            computing_multimedia_start_col_ = 0 
            computing_multimedia_start_row_ += 19
        else:
            computing_multimedia_start_col_ += 5
            computing_multimedia_start_row_ += 0
    
    #alldone        
    elif stringSpecList[0:25]+"..." == "Computing...":
        details.iloc[[0,1]][:].to_excel(writer3, stringSpecList[0:25]+"...", startrow=computing_start_row_, startcol=computing_start_col_, 
                                    header=False, index=False)
        
        print("C")
        
        final_df.to_excel(writer3, stringSpecList[0:25]+"...", startrow=computing_start_row_+3, startcol=computing_start_col_)

        
        computing_count += 1
        if computing_count % 2 == 0:
            computing_start_col_ = 0 
            computing_start_row_ += 19
        else:
            computing_start_col_ += 5
            computing_start_row_ += 0
    
    #alldone        
    elif stringSpecList[0:25]+"..." == "Multimedia...":
        details.iloc[[0,1]][:].to_excel(writer3, stringSpecList[0:25]+"...", startrow=multimedia_start_row_, startcol=multimedia_start_col_, 
                                    header=False, index=False)
        
        print("M")
        final_df.to_excel(writer3, stringSpecList[0:25]+"...", startrow=multimedia_start_row_+3, startcol=multimedia_start_col_)
        
        multimedia_count += 1
        if multimedia_count % 2 == 0:
            multimedia_start_col_ = 0 
            multimedia_start_row_ += 19
        else:
            multimedia_start_col_ += 5
            multimedia_start_row_ += 0
    
    count += 1
    if count % 2 == 0:
        start_col_ = 0 
        start_row_ += 19
    else:
        start_col_ += 5
        start_row_ += 0
writer.save()
#wb2.save('trial.xlsx')
writer2.save()

writer3.save()

#print(df)